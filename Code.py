import os, sys
import numpy as np
import pandas as pd
import yfinance as yfin
import xlwings as xw
import matplotlib.pyplot as plt
import seaborn as sns
from sklearn.linear_model import LinearRegression
from openpyxl import Workbook
from datetime import datetime
from pandas_datareader import data as pdr
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from highlight_text import fig_text
import statsmodels.api as sm

PROJECT_FOLDER_PATH = '_TO_BE_FILLED'

# Line required for pdr.get_data_yahoo() to function properly
yfin.pdr_override()

# Get the tickers of current S&P500 stock list on Wikipedia
stockInfo = pd.read_html('https://en.wikipedia.org/wiki/List_of_S%26P_500_companies')[0]
tickers_np = stockInfo['Symbol'].to_numpy()

# Exclude stocks that do not have full data during analysis period
excludeStocks = np.array(['BF.B', 'BRK.B'])
stockList = [stock for stock in tickers_np if stock not in excludeStocks]

# Start & end dates of in-sample data
startDate_fullAnalysis = '2011-01-01'
endDate_fullAnalysis = '2023-12-31'


"""
Analysis Period         Eval. Year
----------------        ----------
2018-19-20-21-22        2023
2017-18-19-20-21        2022
2016-17-18-19-20        2021
2015-16-17-18-19        2020
2014-15-16-17-18        2019
2013-14-15-16-17        2018
2012-13-14-15-16        2017
2011-12-13-14-15        2016
"""

# 8 analysis periods of 5 years from 2009 to 2022
# For example: range(2018, 2023) means analysis period 2018-19-20-21-22 & evaluation year 2023.
evaluation_year_ranges = {}
for i in range(8): evaluation_year_ranges[2016 + i] = range(2011 + i, 2016 + i)

# Get stocks' adjusted closing prices
# The adjusted closing price includes anything that would affect the stock price (stock splits, dividends...)
def download_stock_prices(stocks: list, start: str, end: str):
    
    stockPricesDf = pdr.get_data_yahoo(stocks, start=start, end=end)
    stockPricesDf = stockPricesDf['Adj Close']

    # Exclude columns that have at least one missing stock price
    stockPricesDf = stockPricesDf.dropna(axis='columns', how='any')

    return stockPricesDf

# Create a dictionary to anonymise stock tickers
def create_anonymisation_mapping(strings):
    
    mapping = {}

    # Generate anonymized names for each string
    for i, string in enumerate(strings):
        anonymized_name = f'stock{i+1}'
        mapping[string] = anonymized_name

    return mapping

# Calculate the risk free rate using Treasury bonds
def calculate_rf_rate_using_treasury(start: str, end: str):
    
    # Download historical data for the Treasury bond: "^IRX" for 13-week Treasury bills
    treasury_symbol = '^IRX'
    treasury_data = yfin.download(treasury_symbol, start=start, end=end)

    # Calculate the average risk-free rate during the analysed period
    risk_free_rate = treasury_data['Close'].mean() / 100.0

    return risk_free_rate

# Export selection from input data between start & end dates
def export_selected_date_range(eval_year: int, stockPricesDf: pd.DataFrame, indexPricesDf: pd.DataFrame, folderPath: str):

    analysis_period = evaluation_year_ranges[eval_year]
    for start_year in analysis_period:

        # Start & end dates of evaluation year
        start = datetime(start_year, 1, 1).strftime('%Y-%m-%d')
        end = datetime(analysis_period[-1], 12, 31).strftime('%Y-%m-%d')
        
        # Extract selection from input data between start & end dates
        stockPrices_partialAnalysis = stockPricesDf.loc[
                        (stockPricesDf.index >= start) & 
                        (stockPricesDf.index <= end)
                    ]
        indexPrices_partialAnalysis = indexPricesDf.loc[
                        (stockPricesDf.index >= start) & 
                        (stockPricesDf.index <= end)
                    ]
        
        # Convert date strings to datetime objects
        start_datetime = datetime.strptime(start, '%Y-%m-%d')
        end_datetime = datetime.strptime(end, '%Y-%m-%d')
        
        # File names
        file_name_stocks = f'{start_datetime.year}_to_{end_datetime.year}_stock_prices_anonymised.csv'
        file_name_index = f'{start_datetime.year}_to_{end_datetime.year}_index_prices.csv'

        # Export as CSV
        stockPrices_partialAnalysis.to_csv(os.path.join(folderPath, file_name_stocks), index=True)
        indexPrices_partialAnalysis.to_csv(os.path.join(folderPath, file_name_index), index=True)

# Write mapping as table into Excel file
def write_mapping_to_excel(excel_file_path: str, mappingDf: pd.DataFrame):

    # Get the name of the first sheet
    xlsFile = pd.ExcelFile(excel_file_path)
    sheet_name = xlsFile.sheet_names[0] if xlsFile.sheet_names else None

    # Add as sheet to Excel report
    with pd.ExcelWriter(excel_file_path, engine='openpyxl', mode='a') as writer:

        # If the sheet exists, remove it
        if sheet_name in writer.book.sheetnames:
            writer.book.remove(writer.book[sheet_name])
        
        # Insert dataframe as new sheet
        mappingDf.to_excel(writer, sheet_name=sheet_name, index=True)

# Calculate return, Alpha & Beta for each stock during the analysed period >> Write to Analysis Excel
def calculate_stock_metrics(eval_year: int, stockPricesDf: pd.DataFrame,
                            stockReturnsDf: pd.DataFrame, indexReturnsDf: pd.DataFrame):
    
    # Start & end dates of evaluation year
    start = datetime(eval_year, 1, 1).strftime('%Y-%m-%d')
    end = datetime(eval_year, 12, 31).strftime('%Y-%m-%d')
    
    # Prepare dataframe for results
    stockResults = stockPricesDf[0:0].copy()
    stockResults = stockResults.rename_axis('Data')

    # Extract selection from input data between start & end dates
    stockPrices_partialAnalysis = stockPricesDf.loc[
                    (stockPricesDf.index >= start) & 
                    (stockPricesDf.index <= end)
                ]
    stockReturns_partialAnalysis = stockReturnsDf.loc[
                    (stockReturnsDf.index >= start) & 
                    (stockReturnsDf.index <= end)
                ]
    indexReturns_partialAnalysis = indexReturnsDf.loc[
                    (indexReturnsDf.index >= start) & 
                    (indexReturnsDf.index <= end)
                ]

    print('\n')
    print('-----------------------------------------------------------')
    print('From', start, 'to', end, ':')

    # Calculate the average risk free rate during the analysed period
    risk_free_rate = calculate_rf_rate_using_treasury(start=start, end=end)
    print('Risk free rate \t\t=', round(100 * risk_free_rate, 2), '%')

    # Calculate the average market return over the analysed period
    average_index_return = indexReturns_partialAnalysis['Index_Returns'].mean() * len(indexReturns_partialAnalysis)
    print('S&P 500 Index Ret.', '\t=', round(100 * average_index_return, 2), '%')
    print('-----------------------------------------------------------')
    print('\n')

    # Merge stock returns & S&P 500 returns (additional safety to ensure that the dates are aligned)
    stocks_index_merged_df = pd.merge(stockReturns_partialAnalysis, indexReturns_partialAnalysis, on='Date', how='inner')

    # Prepare dictionaries for Alpha & Beta
    expected_returns_results = {}
    avg_returns_results = {}
    beta_results = {}
    alpha_results = {}

    # For each stock, calculate return, Alpha & Beta
    for stock in stockPrices_partialAnalysis.columns:

        # Calculate Beta of stock by running a regression analysis.
        # With stock's returns as the dependent variable and the S&P 500's returns as the independent variable.
        # The beta is the slope of the regression line.
        X = stocks_index_merged_df['Index_Returns']
        y = stocks_index_merged_df[stock]
        X1 = sm.add_constant(X)
        stock_model = sm.OLS(y, X1)
        results = stock_model.fit()
        beta = results.params.loc['Index_Returns']

        # Calculate expected return over the analysed period using CAPM
        expected_stock_return = risk_free_rate + beta * (average_index_return - risk_free_rate)

        # Calculate the average stock return over the analysed period
        average_stock_return = stocks_index_merged_df[stock].mean() * len(indexReturns_partialAnalysis)

        # Calculate Alpha
        alpha = average_stock_return - expected_stock_return

        # Add results to dataframe
        expected_returns_results[stock] = expected_stock_return
        avg_returns_results[stock] = average_stock_return
        beta_results[stock] = beta
        alpha_results[stock] = alpha

    # Add results to dataframe    
    stockResults.loc['exp_ret'] = expected_returns_results
    stockResults.loc['avg_ret'] = avg_returns_results
    stockResults.loc['beta'] = beta_results
    stockResults.loc['alpha'] = alpha_results
    
    # Create Excel file if it does not yet exist
    excel_file_path = PROJECT_FOLDER_PATH + f'Analysis - {eval_year}.xlsx'
    if not os.path.exists(excel_file_path):
        workbook = Workbook()
        workbook.save(excel_file_path)

    # Add as sheet to Excel report
    analysis_period = evaluation_year_ranges[eval_year]
    with pd.ExcelWriter(excel_file_path, engine='openpyxl', mode='a') as writer:

        for start_year in analysis_period:

            # Set sheet name
            sheet_name = f'{start_year}_to_{analysis_period[-1]}'

            # If the sheet exists, remove it
            if sheet_name in writer.book.sheetnames:
                writer.book.remove(writer.book[sheet_name])
            
            # Insert dataframe as new sheet
            stockResults.T.to_excel(writer, sheet_name=sheet_name, index=True)

# Predict stock alpha during eval. year using predicted stock prices, predicted index avg return, and real risk free rate
def predict_stock_alpha(stocks_index_merged_returns_Df: pd.DataFrame, stock_lin_reg_models_dict: dict, 
                        predicted_average_index_return: float, eval_risk_free_rate: float, csv_path_str: str):
    
    # Prepare dataframe to be filled then later saved as CSV
    stock_predicted_alpha_df = pd.DataFrame(columns=['Linear_Reg_alpha'])
    stock_predicted_alpha_df.index.name = 'Stock'

    # Calculated predicted Alpha for each stock using stock's linear regression model
    for stock, stock_model in stock_lin_reg_models_dict.items():

        # Predicted Beta is the coefficient of the linear model
        predicted_stock_beta = stock_model.coef_[0]

        # Calculate the predicted stock return
        predicted_average_stock_return = stocks_index_merged_returns_Df[stock].mean() * len(stocks_index_merged_returns_Df)

        # Calculate expected return over the analysed period using CAPM
        expected_stock_return = eval_risk_free_rate + predicted_stock_beta * (predicted_average_index_return - eval_risk_free_rate)

        # Calculate Alpha
        alpha = predicted_average_stock_return - expected_stock_return
        stock_predicted_alpha_df.at[stock, 'Linear_Reg_alpha'] = alpha

    # Export dataframe as CSV
    stock_predicted_alpha_df.to_csv(csv_path_str, index=True)

# Calculate actual & predicted average returns then format text to add below plot figure
def calculate_actual_nd_predicted_avg_return(actual_nd_predicted_prices_Df: pd.DataFrame, col_name_actual: str, col_name_predicted: str, eval_year: int):
                
    actual_returns_Df = actual_nd_predicted_prices_Df[[col_name_actual]].pct_change(fill_method=None).dropna()
    actual_avg_return = actual_returns_Df[col_name_actual].mean() * len(actual_returns_Df)
    
    predicted_returns_Df = actual_nd_predicted_prices_Df[[col_name_predicted]].pct_change(fill_method=None).dropna()
    predicted_avg_return = predicted_returns_Df[col_name_predicted].mean() * len(predicted_returns_Df)

    return f'<Rendement moyen réel en {eval_year}: {round(100 * actual_avg_return, 2)} %>\n<Rendement moyen prédit en {eval_year}: {round(100 * predicted_avg_return, 2)} %>'

# Create & plot a chart with two data series (actual & predicted prices)
def plot_nd_save_dataframe(x_data: pd.Index, transparency_cutoff: str, y_data_1: pd.Series, y_data_2: pd.Series, y_data_1_label: str, 
                           y_data_2_label: str, plot_xlabel: str, plot_ylabel: str, plot_title: str, img_path: str, 
                           txt_below_legend: str = '', chart_info: str = ''):

    fig, ax = plt.subplots(figsize=(12, 6))

    # Plot the first data series
    ax.plot(x_data, y_data_1, label=y_data_1_label, color='#007acc')

    # Split y_data_2 to plot 40% transparent before cutoff
    if transparency_cutoff:
        
        y_data_2_transparent = y_data_2[x_data < transparency_cutoff]
        y_data_2_opaque = y_data_2[x_data >= transparency_cutoff]

        ax.plot(x_data[x_data < transparency_cutoff], y_data_2_transparent, label='_no_legend_', alpha=0.4, color='red', linestyle = 'dotted')
        ax.plot(x_data[x_data >= transparency_cutoff], y_data_2_opaque, label=y_data_2_label, color='red')

    else: ax.plot(x_data, y_data_2, label=y_data_2_label, alpha=0.4, color='red', linestyle = 'dotted')

    ax.set_xlabel(plot_xlabel)
    ax.set_ylabel(plot_ylabel)
    ax.set_title(plot_title)
    ax.legend(loc='upper left')
    ax.set_xticks(x_data)
    ax.set_xticklabels(x_data, rotation=45, fontsize=6)
    ax.xaxis.set_major_locator(plt.MaxNLocator(nbins=30))
    fig.subplots_adjust(bottom=0.2)

    highlight_textprops =\
    [{"bbox": {"edgecolor": "white", "facecolor": "white", "linewidth": 0.7, "pad": 1.5}},
     {"bbox": {"edgecolor": "white", "facecolor": "white", "linewidth": 0.7, "pad": 1.5}}]

    if txt_below_legend:
        fig_text(0.138, 0.76,
                fontsize=8,
                ha='left', va='top',
                s=txt_below_legend,
                highlight_textprops=highlight_textprops,
                ax=ax)
    if chart_info:
        fig_text(0.138, 0.68,
                fontsize=8,
                ha='left', va='top',
                s=chart_info,
                highlight_textprops=highlight_textprops,
                ax=ax)

    plt.savefig(img_path)
    plt.close()


################################################################################################################################################
##### Part 1:   Stock price data
#####           Get prices of each stock + S&P 500 index
################################################################################################################################################
# Get stock prices
stockPricesOriginal = download_stock_prices(stockList, start=startDate_fullAnalysis, end=endDate_fullAnalysis)
stockList = stockPricesOriginal.columns

# Safety check
if stockPricesOriginal.isnull().values.any():
    print(stockPricesOriginal)
    sys.exit('Program interrupted! NaN values found in stock prices!')

# Export original stock prices in CSV
stockPricesOriginal.to_csv(PROJECT_FOLDER_PATH+'stock_prices_original.csv', index=True)

# Create mapping dictionary: original stock name >> anonymised code
# Ex: 'stock1' = mappingDict['AAPL']
mappingDict = create_anonymisation_mapping(stockList)

# Export mapping dictionary as CSV
mappingDf = pd.DataFrame(mappingDict.items(), columns=['Original', 'Anonymised'])
mappingDf.to_csv(PROJECT_FOLDER_PATH+'mapping.csv', index=False)

# Anonymise stock prices
stockPricesAnonymised = stockPricesOriginal.rename(columns=mappingDict, inplace=False)

# Export anonymised stock prices in CSV
stockPricesAnonymised.to_csv(PROJECT_FOLDER_PATH+'stock_prices_anonymised.csv', index=True)

# Download S&P 500 prices
indexPrices = pdr.get_data_yahoo(['^GSPC'], start=startDate_fullAnalysis, end=endDate_fullAnalysis)
indexPrices = indexPrices['Adj Close']

# Export S&P 500 prices as CSV
indexPrices.to_csv(PROJECT_FOLDER_PATH+'index_prices.csv', index=True)


################################################################################################################################################
##### Part 2:   Alpha
#####           For each stock: calculate alpha during eval. year
################################################################################################################################################

# Load mapping data
mappingDf = pd.read_csv(PROJECT_FOLDER_PATH+'mapping.csv', index_col=0)

# Create Excel report files if do not yet exist
for eval_year in evaluation_year_ranges:
    excel_file_path = PROJECT_FOLDER_PATH + f'Analysis - {eval_year}.xlsx'
    if not os.path.exists(excel_file_path):
        workbook = Workbook()
        workbook.save(excel_file_path)

    # Write mapping to first sheet of Excel
    write_mapping_to_excel(excel_file_path, mappingDf)

# Load S&P 500 Index data
indexPrices = pd.read_csv(PROJECT_FOLDER_PATH+'index_prices.csv', index_col=0)

# Load original stock prices
stockPricesOriginal = pd.read_csv(PROJECT_FOLDER_PATH+'stock_prices_original.csv', index_col=0)

# Calculate the stock returns & drop NaN values
stockReturns = stockPricesOriginal.pct_change(fill_method=None).dropna()

# Calculate the index returns & drop NaN values
indexReturns = indexPrices.pct_change(fill_method=None).dropna()
indexReturns.rename(columns={'Adj Close': 'Index_Returns'}, inplace=True)

# For each stock, calculate return, Alpha & Beta during each evaluation year (8 years, one for each analysis period)
for eval_year in evaluation_year_ranges: calculate_stock_metrics(eval_year, stockPricesOriginal, stockReturns, indexReturns)


################################################################################################################################################
##### Part 3:   Export partial prices (index & stocks)
################################################################################################################################################

# Load anonymised stock prices
stockPricesAnonymised = pd.read_csv(PROJECT_FOLDER_PATH+'stock_prices_anonymised.csv', index_col=0)

# Create folders for partial prices CSV files if do not yet exist
for eval_year in evaluation_year_ranges:
    partial_res_folder_path = PROJECT_FOLDER_PATH + f'Partial_Prices_{eval_year}'
    if not os.path.exists(partial_res_folder_path): os.makedirs(partial_res_folder_path)

# Export partial index prices & anonymised stock prices of each analysis period as CSV
for eval_year in evaluation_year_ranges:
    partial_res_folder_path = PROJECT_FOLDER_PATH + f'Partial_Prices_{eval_year}'
    export_selected_date_range(eval_year, stockPricesAnonymised, indexPrices, partial_res_folder_path)


################################################################################################################################################
##### Part 4:   Build linear regression model & add export results as CSV files
################################################################################################################################################

# Load S&P 500 Index data
indexPrices = pd.read_csv(PROJECT_FOLDER_PATH+'index_prices.csv', index_col=0)
indexPrices.rename(columns={'Adj Close': 'Index_Prices'}, inplace=True)

# Load original stock prices
stockPricesOriginal = pd.read_csv(PROJECT_FOLDER_PATH+'stock_prices_original.csv', index_col=0)

# Merge stock prices & S&P 500 prices (additional safety to ensure that the dates are aligned)
stocks_index_merged_prices_df = pd.merge(stockPricesOriginal, indexPrices, on='Date', how='inner')

# For each eval year >> an Excel file >> multiple sheets.
# In each sheet, calculate metrics for eval. year based on sheet range.
for eval_year in evaluation_year_ranges:

    # Start & end dates of evaluation year
    eval_start = datetime(eval_year, 1, 1).strftime('%Y-%m-%d')
    eval_end = datetime(eval_year, 12, 31).strftime('%Y-%m-%d')

    # Calculate the actual average risk free rate during evaluation year
    eval_risk_free_rate = calculate_rf_rate_using_treasury(start=eval_start, end=eval_end)
    print(f'Risk free rate from {eval_start} to {eval_end} =', round(100 * eval_risk_free_rate, 2), '%')

    analysis_period = evaluation_year_ranges[eval_year]
    lin_reg_folder_path = os.path.join(PROJECT_FOLDER_PATH, f'Lin_Reg_predictions_{eval_year}')
    
    # Create subfolders for plots (eval. year)
    lin_reg_plot_training_folder = os.path.join(lin_reg_folder_path, '___plots_training')
    lin_reg_plot_full_folder = os.path.join(lin_reg_folder_path, '___plots_full')
    if not os.path.exists(lin_reg_plot_training_folder): os.makedirs(lin_reg_plot_training_folder)
    if not os.path.exists(lin_reg_plot_full_folder): os.makedirs(lin_reg_plot_full_folder)

    for training_start_year in analysis_period:

        # Extract selection from input data between start & end dates
        stocks_index_merged_prices_evalYear = stocks_index_merged_prices_df.loc[
                    (stocks_index_merged_prices_df.index >= eval_start) & 
                    (stocks_index_merged_prices_df.index <= eval_end)
                ]
        
        # Data Processing and scaling. Reset index and convert it to column
        stocks_index_merged_prices_evalYear.reset_index(inplace=True)

        training_end_year = analysis_period[-1]

        # Start & end dates of input data used to train linear regression model
        training_start = datetime(training_start_year, 1, 1).strftime('%Y-%m-%d')
        training_end = datetime(training_end_year, 12, 31).strftime('%Y-%m-%d')
        
        print('-----------------------------------------------------------')
        print('From', training_start, 'to', training_end, ':')
        print(f'Data from {training_start_year} to {training_end_year} used to predict Alpha during {eval_year}')
        print('--------------------------------------------------------------------------------')

        # Create subfolders for plots (analysis period)
        lin_reg_plot_training_folder_path = os.path.join(lin_reg_plot_training_folder, f'{training_start_year}_{training_end_year}')
        lin_reg_plot_full_folder_path = os.path.join(lin_reg_plot_full_folder, f'{training_start_year}_{training_end_year}')
        if not os.path.exists(lin_reg_plot_training_folder_path): os.makedirs(lin_reg_plot_training_folder_path)
        if not os.path.exists(lin_reg_plot_full_folder_path): os.makedirs(lin_reg_plot_full_folder_path)

        # Extract selection from input data between start & end dates
        stocks_index_merged_prices_partialAnalysis = stocks_index_merged_prices_df.loc[
                    (stocks_index_merged_prices_df.index >= training_start) & 
                    (stocks_index_merged_prices_df.index <= training_end)
                ]
        
        # Data Processing and scaling. Reset index and convert it to column
        stocks_index_merged_prices_partialAnalysis.reset_index(inplace=True)

        ######### Create lin. reg. model (index)
        # Reshape index column to 2D array for .fit() method (index)
        X_train = np.array(stocks_index_merged_prices_partialAnalysis.index).reshape(-1, 1)
        y_train = stocks_index_merged_prices_partialAnalysis['Index_Prices']
        index_lin_reg_model = LinearRegression()
        index_lin_reg_model.fit(X_train, y_train)

        # Apply lin. reg. model on training data (index)
        predicted_y_train = index_lin_reg_model.predict(X_train)
        predicted_y_train_df = pd.DataFrame(predicted_y_train, index=y_train.index, columns=['predicted_price'])

        # Apply flooring to replace predicted prices below 1$ with 1$
        predicted_y_train_df['predicted_price'] = predicted_y_train_df['predicted_price'].apply(lambda x: max(x, 1))

        # Plot actual & predicted prices for training data (index)
        model_applied_to_training_df = pd.concat([y_train, predicted_y_train_df], axis=1).rename(columns={'Index_Prices': 'actual_price'}, inplace=False).set_index(pd.Index(stocks_index_merged_prices_partialAnalysis['Date']), inplace=False)
        plot_nd_save_dataframe(model_applied_to_training_df.index, 0, model_applied_to_training_df['actual_price'], model_applied_to_training_df['predicted_price'], 'Cours réels', 'Cours prédits', 
                                'Jours', 'Cours', "Cours réels et prédits de l'indice S&P 500", os.path.join(lin_reg_plot_training_folder_path, '____Index.png'),
                                f"<Modèle entrainé puis appliqué entre {training_start_year} et {training_end_year}><>")
        
        # Apply index lin. reg. model on eval. year & add to overall df (index)
        stocks_index_merged_prices_evalYear.index = stocks_index_merged_prices_evalYear.index + len(stocks_index_merged_prices_partialAnalysis)
        X_evalYear = np.array(stocks_index_merged_prices_evalYear.index).reshape(-1, 1)
        predicted_index_evalYear = index_lin_reg_model.predict(X_evalYear)

        # Prepare df for predicted prices during eval. year (to be used for index & stocks)
        predicted_stocks_index_prices_evalYear_df = pd.DataFrame(data=[], index=stocks_index_merged_prices_evalYear.index, columns=stocks_index_merged_prices_evalYear.columns)
        predicted_stocks_index_prices_evalYear_df.rename(columns={'Index_Prices': 'Index'}, inplace=True)
        predicted_stocks_index_prices_evalYear_df.drop(columns=['Date'], inplace=True)
        
        # Prepare actual & predicted prices for eval. year (index)
        predicted_stocks_index_prices_evalYear_df['Index'] = pd.Series(np.array(predicted_index_evalYear), name='Index', index=predicted_stocks_index_prices_evalYear_df.index)
        model_applied_to_evalYear_df = pd.concat([stocks_index_merged_prices_evalYear['Index_Prices'], predicted_stocks_index_prices_evalYear_df['Index']], axis=1).rename(columns={'Index_Prices': 'actual_price', 'Index': 'predicted_price'}, inplace=False).set_index(pd.Index(stocks_index_merged_prices_evalYear['Date']), inplace=False)

        # Print characteristics (intercept & coef.) & performance (RMSE, R-Squ...) of lin. reg. model of index when applied to eval. year
        print('.')
        print('--------------------------------')
        index_coefficient = index_lin_reg_model.coef_[0]
        index_intercept = index_lin_reg_model.intercept_
        index_r_squared = index_lin_reg_model.score(X_train, y_train)
        diff=np.subtract(model_applied_to_evalYear_df['actual_price'], model_applied_to_evalYear_df['predicted_price'])
        square=np.square(diff)
        MSE=square.mean()
        index_rmse=np.sqrt(MSE)
        print("Ordonnée à l'origine : \t{:.2f}".format(index_intercept))
        print("Pente : \t\t{:.2f}".format(index_coefficient))
        print("R-carré : \t\t{:.3f}".format(index_r_squared))
        print("REQM : \t\t\t{:.2f}".format(index_rmse))
        print('--------------------------------')
        print('.')
        
        # Apply flooring to replace predicted prices below 1$ with 1$
        predicted_stocks_index_prices_evalYear_df['Index'] = predicted_stocks_index_prices_evalYear_df['Index'].apply(lambda x: max(x, 1))

        # Calculate predicted index return during eval. year for Alpha calculations (index)
        predicted_returns_df = model_applied_to_evalYear_df[['predicted_price']].pct_change(fill_method=None).dropna()
        predicted_eval_average_index_return = predicted_returns_df['predicted_price'].mean() * len(predicted_returns_df)
        
        # Predict & plot prices during [training data + eval. year] (index)
        train_plus_evalYear = pd.concat([model_applied_to_training_df, model_applied_to_evalYear_df], ignore_index=False)
        index_transp_cutoff = model_applied_to_training_df.index[-1]
        text_below_legend = calculate_actual_nd_predicted_avg_return(model_applied_to_evalYear_df, 'actual_price', 'predicted_price', eval_year)
        plot_nd_save_dataframe(train_plus_evalYear.index, index_transp_cutoff, train_plus_evalYear['actual_price'], 
                                train_plus_evalYear['predicted_price'], 'Cours réels', 'Cours prédits', 
                                'Jours', 'Cours', "Cours réels et prédits de l'indice S&P 500", os.path.join(lin_reg_plot_full_folder_path, '____Index.png'),
                                text_below_legend, f"<Modèle entrainé entre {training_start_year} et {training_end_year}>\n<puis appliqué sur {eval_year}>")
        
        # Create & apply lin. reg. model for each stock
        stock_lin_reg_models = {}
        for stock in stockPricesOriginal.columns:

            ######### Create lin. reg. model (stock)
            # Reshape index column to 2D array for .fit() method (stock)
            X_train_stock = np.array(stocks_index_merged_prices_partialAnalysis.index).reshape(-1, 1)
            y_train_stock = stocks_index_merged_prices_partialAnalysis[stock]
            stock_lin_reg_model = LinearRegression()
            stock_lin_reg_model.fit(X_train_stock, y_train_stock)
            stock_lin_reg_models[stock] = stock_lin_reg_model
            
            # Apply lin. reg. model on training data (stock)
            predicted_y_train = stock_lin_reg_model.predict(X_train_stock)
            predicted_y_train_df = pd.DataFrame(predicted_y_train, index=y_train_stock.index, columns=['predicted_price'])

            # Apply flooring to replace predicted prices below 1$ with 1$
            predicted_y_train_df['predicted_price'] = predicted_y_train_df['predicted_price'].apply(lambda x: max(x, 1))
            
            # Plot actual & predicted prices for training data (stock)
            model_applied_to_training_df = pd.concat([y_train_stock, predicted_y_train_df], axis=1).rename(columns={stock: 'actual_price'}, inplace=False).set_index(pd.Index(stocks_index_merged_prices_partialAnalysis['Date']), inplace=False)
            plot_nd_save_dataframe(model_applied_to_training_df.index, 0, model_applied_to_training_df['actual_price'], 
                                    model_applied_to_training_df['predicted_price'], 'Cours réels', 'Cours prédits', 'Jours', 'Cours', 
                                    f"Cours réels et prédits de l'action ({stock})", os.path.join(lin_reg_plot_training_folder_path, f'___{stock}.png'), 
                                    f"<Modèle entrainé puis appliqué entre {training_start_year} et {training_end_year}><>")
            
            # Apply stock lin. reg. model on eval. year & add to overall df (stock)
            predicted_stock_evalYear = stock_lin_reg_model.predict(X_evalYear)
            predicted_stocks_index_prices_evalYear_df[stock] = pd.Series(np.array(predicted_stock_evalYear), name=stock, index=predicted_stocks_index_prices_evalYear_df.index)
            
            # Apply flooring to replace predicted prices below 1$ with 1$
            predicted_stocks_index_prices_evalYear_df[stock] = predicted_stocks_index_prices_evalYear_df[stock].apply(lambda x: max(x, 1))

            # Predict & plot actual & predicted prices for eval. year (stock)
            model_applied_to_evalYear_df = pd.concat([stocks_index_merged_prices_evalYear[stock], predicted_stocks_index_prices_evalYear_df[stock]], axis=1).set_index(pd.Index(stocks_index_merged_prices_evalYear['Date']), inplace=False)
            model_applied_to_evalYear_df.columns = ['actual_price', 'predicted_price']
            
            # Predict & plot prices during [training data + eval. year] (stock)
            train_plus_evalYear = pd.concat([model_applied_to_training_df, model_applied_to_evalYear_df], ignore_index=False)
            stock_transp_cutoff = model_applied_to_training_df.index[-1]
            text_below_legend = calculate_actual_nd_predicted_avg_return(model_applied_to_evalYear_df, 'actual_price', 'predicted_price', eval_year)
            plot_nd_save_dataframe(train_plus_evalYear.index, stock_transp_cutoff, train_plus_evalYear['actual_price'], 
                                    train_plus_evalYear['predicted_price'], 'Cours réels', 'Cours prédits', 'Jours', 'Cours', 
                                    f"Cours réels et prédits de l'action ({stock})", os.path.join(lin_reg_plot_full_folder_path, f'___{stock}.png'), 
                                    text_below_legend, f"<Modèle entrainé entre {training_start_year} et {training_end_year}>\n<puis appliqué sur {eval_year}>")

        # Calculate the stock returns & drop NaN values
        predicted_stocks_index_returns_evalYear_df = predicted_stocks_index_prices_evalYear_df.pct_change(fill_method=None).dropna()
        
        # For each stock, calculate predicted Alpha then write CSV
        csv_path = os.path.join(lin_reg_folder_path, f'{training_start_year}_to_{training_end_year}_alpha_predictions.csv')
        print(f'Eval. year = {eval_year}: Training from {training_start_year} to {training_end_year}.')
        predict_stock_alpha(predicted_stocks_index_returns_evalYear_df, stock_lin_reg_models, predicted_eval_average_index_return, eval_risk_free_rate, csv_path)

        

################################################################################################################################################
##### Part 5:   Import predictions of ChatGPT & Lin. Reg. model from CSV files to Excel files
################################################################################################################################################

# Data will be inserted in a range bottom left starting from this cell
start_row = 1
start_col_chatgpt = 9
start_col_lin_reg = 18

# For each eval year >> an Excel file. In each file, import the relevant CSV to the relevant sheet.
for eval_year in evaluation_year_ranges:
    
    excel_file_path = PROJECT_FOLDER_PATH + f'Analysis - {eval_year}.xlsx'
    chatgpt_folder_path = os.path.join(PROJECT_FOLDER_PATH, f'ChatGPT_predictions_{eval_year}')
    lin_reg_folder_path = os.path.join(PROJECT_FOLDER_PATH, f'Lin_Reg_predictions_{eval_year}')

    try:
        # Load the existing Excel file
        book = load_workbook(excel_file_path)

        # Each Excel file has multiple sheets
        analysis_period = evaluation_year_ranges[eval_year]
        for start_year in analysis_period:

            # Sheet name
            sheet_name = f'{start_year}_to_{analysis_period[-1]}'

            # Select the sheet
            sheet = book[sheet_name]

            # CSV file path
            chatgpt_file_path = os.path.join(chatgpt_folder_path, f'{sheet_name}_alpha_predictions.csv')
            lin_reg_file_path = os.path.join(lin_reg_folder_path, f'{sheet_name}_alpha_predictions.csv')

            # Import CSV file content as DataFrame
            chatgpt_csv_df = pd.read_csv(chatgpt_file_path, index_col=0)
            lin_reg_csv_df = pd.read_csv(lin_reg_file_path, index_col=0)

            ################### ChatGPT
            # Write the ChatGPT DataFrame to the sheet starting from the next empty row
            next_row = start_row
            for row in dataframe_to_rows(chatgpt_csv_df, index=True, header=False):
                for col_idx, value in enumerate(row, start_col_chatgpt):
                    sheet.cell(row=next_row, column=col_idx, value=value)
                next_row += 1

            # Safety signature
            sheet.cell(row=next_row+2, column=start_col_chatgpt, 
                        value=f'Imported from ChatGPT file {sheet_name}_alpha_predictions.csv to {sheet_name}.')

            ################### Lin. Reg.
            # Write the Lin. Reg. DataFrame to the sheet starting from the next empty row
            next_row = start_row
            for row in dataframe_to_rows(lin_reg_csv_df, index=True, header=False):
                for col_idx, value in enumerate(row, start_col_lin_reg):
                    sheet.cell(row=next_row, column=col_idx, value=value)
                next_row += 1

            # Safety signature
            sheet.cell(row=next_row+2, column=start_col_lin_reg, 
                        value=f'Imported from Lin. Reg. file {sheet_name}_alpha_predictions.csv to {sheet_name}.')

            # Save the changes to the Excel file
            book.save(excel_file_path)

    except Exception as e:
        print(f'Error: {e}')
        sys.exit(f'Failed to write data into Excel file: Analysis - {eval_year}.xlsx!')


################################################################################################################################################
##### Part 6:   Collect results (accuracy, RMSE, R-squared) from Excel files
################################################################################################################################################

# Recalculate all formulas in Excel files
app = xw.App(visible=False)
try:
    for eval_year in evaluation_year_ranges:
    
        excel_file_path = PROJECT_FOLDER_PATH + f'Analysis - {eval_year}.xlsx'
        workbook = app.books.open(excel_file_path)
        workbook.app.calculate()
        workbook.save()

finally: app.quit()

# Prepare 2 dfs for the results of each model
results_chatgpt_df = pd.DataFrame(data=[], columns=['Sign_F1_score', 'Sign_Acc', 'Value_RMSE', 'Value_R_Squ', 'Eval_Year', 'Nb_Years_Train'])
results_lin_reg_df = pd.DataFrame(data=[], columns=['Sign_F1_score', 'Sign_Acc', 'Value_RMSE', 'Value_R_Squ', 'Eval_Year', 'Nb_Years_Train'])

# Populating the DataFrame with ranges as index grouped by eval_year
for eval_year in evaluation_year_ranges:
    
    excel_file_path = PROJECT_FOLDER_PATH + f'Analysis - {eval_year}.xlsx'
    analysis_period = evaluation_year_ranges[eval_year]

    # Read the data from the Excel files
    try:
        # Load the existing Excel file
        workbook = load_workbook(excel_file_path, data_only=True)

        for start_year in analysis_period:

            # Sheet name
            sheet_name = f'{start_year}_to_{analysis_period[-1]}'
            if sheet_name in workbook.sheetnames:

                # Select the sheet
                sheet = workbook[sheet_name]

                # Number of years in training interval
                nb_years_training = analysis_period[-1] - start_year + 1

                # Get ChatGPT results from cells
                r1 = sheet['P8'].value
                r2 = sheet['P9'].value
                r3 = sheet['P12'].value
                r4 = sheet['P13'].value

                # Insert results in ChatGPT df
                results_chatgpt_df.loc[sheet_name] = {'Sign_F1_score': r1, 'Sign_Acc': r2, 'Value_RMSE': r3, 
                                                      'Value_R_Squ': r4, 'Eval_Year': eval_year, 'Nb_Years_Train': nb_years_training}

                # Get Lin. Reg. results from cells
                x1 = sheet['V8'].value
                x2 = sheet['V9'].value
                x3 = sheet['V12'].value
                x4 = sheet['V13'].value

                # Insert results in Lin. Reg. df
                results_lin_reg_df.loc[sheet_name] = {'Sign_F1_score': x1, 'Sign_Acc': x2, 'Value_RMSE': x3, 
                                                      'Value_R_Squ': x4, 'Eval_Year': eval_year, 'Nb_Years_Train': nb_years_training}
                
            else: sys.exit(f'Sheet {sheet_name} not found in Analysis - {eval_year}.xlsx!')

        # Close the workbook
        workbook.close()

    except Exception as e:
        print(f'Error: {e}')
        sys.exit('Failed to load data from at least one of the Excel files!')

# Set the style of seaborn
sns.set(style="whitegrid")

# Create a scatter plot
plt.figure(figsize=(12, 8))
plt.close()
plt.clf()

###### CHART 1a
# Group by 'Nb_Years_Train' and calculate the average 'Value_R_Squ' for each 'Nb_Years_Train'
avg_r_squ_chatgpt = results_chatgpt_df.groupby('Nb_Years_Train')['Value_R_Squ'].mean()
avg_r_squ_lin_reg = results_lin_reg_df.groupby('Nb_Years_Train')['Value_R_Squ'].mean()

plt.scatter(avg_r_squ_chatgpt.index, avg_r_squ_chatgpt.values, label='ChatGPT', color='green', marker='X', s=50)
plt.scatter(avg_r_squ_lin_reg.index, avg_r_squ_lin_reg.values, label='Régression linéaire', color='blue', marker='X', s=50)

plt.ylim(top=1.2*max(avg_r_squ_chatgpt.max(), avg_r_squ_lin_reg.max()))
plt.xticks(results_chatgpt_df['Nb_Years_Train'].unique())

plt.title("R-carré moyen sur les périodes d'entrainement (de 1 à 5 ans)")
plt.legend(loc='upper left')
plt.xlabel("Largeur d'intervalle d'entrainement (années)")
plt.ylabel("R-carré moyen")

# Fit a polynomial (trend line) using numpy.polyfit, adjust degree as needed
degree = 2
coefficients_df1 = np.polyfit(avg_r_squ_chatgpt.index, avg_r_squ_chatgpt.values, degree)
coefficients_df2 = np.polyfit(avg_r_squ_lin_reg.index, avg_r_squ_lin_reg.values, degree)

# Create the trend line using numpy.poly1d
trend_line_df1 = np.poly1d(coefficients_df1)
trend_line_df2 = np.poly1d(coefficients_df2)

# Plot the trend line
plt.plot(avg_r_squ_chatgpt.index, trend_line_df1(avg_r_squ_chatgpt.index), color='green', linestyle='-', alpha=0.3)
plt.plot(avg_r_squ_lin_reg.index, trend_line_df2(avg_r_squ_lin_reg.index), color='blue', linestyle='-', alpha=0.3)
plt.savefig('1a_Avg_R_Squ_Across_Eval_Years.png', dpi=300)
plt.close()
plt.clf()

###### CHART 1b
# Group by 'Eval_Year' and calculate the average 'Value_R_Squ' for each 'Eval_Year'
avg_r_squ_chatgpt = results_chatgpt_df.groupby('Eval_Year')['Value_R_Squ'].mean()
avg_r_squ_lin_reg = results_lin_reg_df.groupby('Eval_Year')['Value_R_Squ'].mean()

plt.scatter(avg_r_squ_chatgpt.index, avg_r_squ_chatgpt.values, label='ChatGPT', s=110, c='white', edgecolors='green', linewidth=1.5)
plt.scatter(avg_r_squ_lin_reg.index, avg_r_squ_lin_reg.values, label='Régression linéaire', s=50, c='blue', edgecolors='blue', linewidth=0)

plt.ylim(top=1.2*max(avg_r_squ_chatgpt.max(), avg_r_squ_lin_reg.max()))
plt.title("R-carré moyen sur chaque année d'évaluation")
plt.legend(loc='upper right')
plt.xlabel("Année d'évaluation")
plt.ylabel("R-carré moyen")
plt.savefig('1b_Avg_R_Squ_Across_Eval_Years.png', dpi=300)
plt.close()
plt.clf()

###### CHART 1c
# Scatter of 'Value_R_Squ' from each experiment of each 'Eval_Year'
merged_df = pd.merge(results_chatgpt_df, results_lin_reg_df, left_index=True, right_index=True, suffixes=('_df1', '_df2'))
plt.scatter(merged_df['Eval_Year_df1'], merged_df['Value_R_Squ_df1'], label='ChatGPT', s=100, c='white', edgecolors='green', linewidth=2)
plt.scatter(merged_df['Eval_Year_df2'], merged_df['Value_R_Squ_df2'], label='Régression linéaire', s=50, c='blue', edgecolors='blue', linewidth=0)

plt.ylim(bottom=-0.02, top=1.4*max(merged_df['Value_R_Squ_df1'].max(), merged_df['Value_R_Squ_df2'].max()))
plt.xlabel("Année d'évaluation")
plt.ylabel("R-carré")
plt.title("R-carré de chaque instance de test (5 tests par année d'évaluation)")
plt.legend(loc='upper left')
plt.savefig('1c_R_Squ_of_each_test.png', dpi=300)
plt.close()
plt.clf()
#------------------------------------------------------------------------------------

###### CHART 2a
# Group by 'Nb_Years_Train' and calculate the average 'Value_RMSE' for each 'Nb_Years_Train'
avg_r_squ_chatgpt = results_chatgpt_df.groupby('Nb_Years_Train')['Value_RMSE'].mean()
avg_r_squ_lin_reg = results_lin_reg_df.groupby('Nb_Years_Train')['Value_RMSE'].mean()

plt.scatter(avg_r_squ_chatgpt.index, avg_r_squ_chatgpt.values, label='ChatGPT', color='green', marker='X', s=50)
plt.scatter(avg_r_squ_lin_reg.index, avg_r_squ_lin_reg.values, label='Régression linéaire', color='blue', marker='X', s=50)

plt.ylim(bottom=0.6*min(avg_r_squ_chatgpt.min(), avg_r_squ_lin_reg.min()), top=1.2*max(avg_r_squ_chatgpt.max(), avg_r_squ_lin_reg.max()))
plt.xticks(results_chatgpt_df['Nb_Years_Train'].unique())

plt.title("REQM moyenne sur les périodes d'entrainement (de 1 à 5 ans)")
plt.legend(loc='upper right')
plt.xlabel("Largeur d'intervalle d'entrainement (années)")
plt.ylabel('REQM moyenne')

# Fit a polynomial (trend line) using numpy.polyfit, adjust degree as needed
degree = 2
coefficients_df1 = np.polyfit(avg_r_squ_chatgpt.index, avg_r_squ_chatgpt.values, degree)
coefficients_df2 = np.polyfit(avg_r_squ_lin_reg.index, avg_r_squ_lin_reg.values, degree)

# Create the trend line using numpy.poly1d
trend_line_df1 = np.poly1d(coefficients_df1)
trend_line_df2 = np.poly1d(coefficients_df2)

# Plot the trend line
plt.plot(avg_r_squ_chatgpt.index, trend_line_df1(avg_r_squ_chatgpt.index), color='green', linestyle='-', alpha=0.3)
plt.plot(avg_r_squ_lin_reg.index, trend_line_df2(avg_r_squ_lin_reg.index), color='blue', linestyle='-', alpha=0.3)
plt.savefig('2a_Avg_RMSE_Across_Eval_Years.png', dpi=300)
plt.close()
plt.clf()

###### CHART 2b
# Group by 'Eval_Year' and calculate the average 'Value_RMSE' for each 'Eval_Year'
avg_r_squ_chatgpt = results_chatgpt_df.groupby('Eval_Year')['Value_RMSE'].mean()
avg_r_squ_lin_reg = results_lin_reg_df.groupby('Eval_Year')['Value_RMSE'].mean()

plt.scatter(avg_r_squ_chatgpt.index, avg_r_squ_chatgpt.values, label='ChatGPT', color='green', marker='X', s=50)
plt.scatter(avg_r_squ_lin_reg.index, avg_r_squ_lin_reg.values, label='Régression linéaire', color='blue', marker='X', s=50)

plt.ylim(bottom=0.8*min(avg_r_squ_chatgpt.min(), avg_r_squ_lin_reg.min()), top=1.2*max(avg_r_squ_chatgpt.max(), avg_r_squ_lin_reg.max()))
plt.title("REQM moyenne sur chaque année d'évaluation")
plt.legend(loc='upper left')
plt.xlabel("Année d'évaluation")
plt.ylabel('REQM moyenne')
plt.savefig('2b_Avg_RMSE_Across_Eval_Years.png', dpi=300)
plt.close()
plt.clf()

###### CHART 2c
# Scatter of 'Value_RMSE' from each experiment of each 'Eval_Year'
merged_df = pd.merge(results_chatgpt_df, results_lin_reg_df, left_index=True, right_index=True, suffixes=('_df1', '_df2'))
plt.scatter(merged_df['Eval_Year_df1'], merged_df['Value_RMSE_df1'], label='ChatGPT', s=100, c='white', edgecolors='green', linewidth=1)
plt.scatter(merged_df['Eval_Year_df2'], merged_df['Value_RMSE_df2'], label='Régression linéaire', s=50, c='blue', edgecolors='blue', linewidth=0)

plt.ylim(bottom=0, top=1.2*max(merged_df['Value_RMSE_df1'].max(), merged_df['Value_RMSE_df2'].max()))
plt.xlabel("Année d'évaluation")
plt.ylabel('REQM')
plt.title("REQM de chaque instance de test (5 tests par année d'évaluation)")
plt.legend(loc='upper left')
plt.savefig('2c_RMSE_of_each_test.png', dpi=300)
plt.close()
plt.clf()
#------------------------------------------------------------------------------------

###### CHART 3a
# Group by 'Nb_Years_Train' and calculate the average 'Sign_Acc' for each 'Nb_Years_Train'
avg_r_squ_chatgpt = results_chatgpt_df.groupby('Nb_Years_Train')['Sign_Acc'].mean()
avg_r_squ_lin_reg = results_lin_reg_df.groupby('Nb_Years_Train')['Sign_Acc'].mean()

plt.scatter(avg_r_squ_chatgpt.index, avg_r_squ_chatgpt.values, label='ChatGPT', color='green', marker='X', s=50)
plt.scatter(avg_r_squ_lin_reg.index, avg_r_squ_lin_reg.values, label='Régression linéaire', color='blue', marker='X', s=50)

plt.ylim(bottom=0.8*min(avg_r_squ_chatgpt.min(), avg_r_squ_lin_reg.min()), top=1.2*max(avg_r_squ_chatgpt.max(), avg_r_squ_lin_reg.max()))
plt.xticks(results_chatgpt_df['Nb_Years_Train'].unique())

plt.title("Exactitude moyenne du signe d'Alpha sur les périodes d'entrainement (de 1 à 5 ans)")
plt.legend(loc='upper right')
plt.xlabel("Largeur d'intervalle d'entrainement (années)")
plt.ylabel("Exactitude moyenne du signe d'Alpha")

# Fit a polynomial (trend line) using numpy.polyfit, adjust degree as needed
degree = 2
coefficients_df1 = np.polyfit(avg_r_squ_chatgpt.index, avg_r_squ_chatgpt.values, degree)
coefficients_df2 = np.polyfit(avg_r_squ_lin_reg.index, avg_r_squ_lin_reg.values, degree)

# Create the trend line using numpy.poly1d
trend_line_df1 = np.poly1d(coefficients_df1)
trend_line_df2 = np.poly1d(coefficients_df2)

# Plot the trend line
plt.plot(avg_r_squ_chatgpt.index, trend_line_df1(avg_r_squ_chatgpt.index), color='green', linestyle='-', alpha=0.3)
plt.plot(avg_r_squ_lin_reg.index, trend_line_df2(avg_r_squ_lin_reg.index), color='blue', linestyle='-', alpha=0.3)
plt.savefig('3a_Avg_Sign_Acc_Across_Eval_Years.png', dpi=300)
plt.close()
plt.clf()

###### CHART 3b
# Group by 'Eval_Year' and calculate the average 'Sign_Acc' for each 'Eval_Year'
avg_r_squ_chatgpt = results_chatgpt_df.groupby('Eval_Year')['Sign_Acc'].mean()
avg_r_squ_lin_reg = results_lin_reg_df.groupby('Eval_Year')['Sign_Acc'].mean()

plt.scatter(avg_r_squ_chatgpt.index, avg_r_squ_chatgpt.values, label='ChatGPT', color='green', marker='X', s=50)
plt.scatter(avg_r_squ_lin_reg.index, avg_r_squ_lin_reg.values, label='Régression linéaire', color='blue', marker='X', s=50)

plt.ylim(bottom=0.8*min(avg_r_squ_chatgpt.min(), avg_r_squ_lin_reg.min()), top=1.2*max(avg_r_squ_chatgpt.max(), avg_r_squ_lin_reg.max()))
plt.title("Exactitude moyenne du signe d'Alpha sur chaque année d'évaluation")
plt.legend(loc='upper left')
plt.xlabel("Année d'évaluation")
plt.ylabel("Exactitude moyenne du signe d'Alpha")
plt.savefig('3b_Avg_Sign_Acc_Across_Eval_Years.png', dpi=300)
plt.close()
plt.clf()

###### CHART 3c
# Scatter of 'Sign_Acc' from each experiment of each 'Eval_Year'
merged_df = pd.merge(results_chatgpt_df, results_lin_reg_df, left_index=True, right_index=True, suffixes=('_df1', '_df2'))
plt.scatter(merged_df['Eval_Year_df1'], merged_df['Sign_Acc_df1'], label='ChatGPT', s=100, c='white', edgecolors='green', linewidth=2)
plt.scatter(merged_df['Eval_Year_df2'], merged_df['Sign_Acc_df2'], label='Régression linéaire', s=50, c='blue', edgecolors='blue', linewidth=0)

plt.ylim(bottom=0.8*min(merged_df['Sign_Acc_df1'].min(), merged_df['Sign_Acc_df2'].min()), top=1.2*max(merged_df['Sign_Acc_df1'].max(), merged_df['Sign_Acc_df2'].max()))
plt.xlabel("Année d'évaluation")
plt.ylabel("Exactitude du signe d'Alpha")
plt.title("Exactitude du signe d'Alpha pour chaque instance de test (5 tests par année d'évaluation)", fontsize = 11.5)
plt.legend(loc='upper left')
plt.savefig('3c_Sign_Acc_of_each_test.png', dpi=300)
plt.close()
plt.clf()
#------------------------------------------------------------------------------------

###### CHART 4a
# Group by 'Nb_Years_Train' and calculate the average 'Sign_F1_score' for each 'Nb_Years_Train'
avg_r_squ_chatgpt = results_chatgpt_df.groupby('Nb_Years_Train')['Sign_F1_score'].mean()
avg_r_squ_lin_reg = results_lin_reg_df.groupby('Nb_Years_Train')['Sign_F1_score'].mean()

plt.scatter(avg_r_squ_chatgpt.index, avg_r_squ_chatgpt.values, label='ChatGPT', color='green', marker='X', s=50)
plt.scatter(avg_r_squ_lin_reg.index, avg_r_squ_lin_reg.values, label='Régression linéaire', color='blue', marker='X', s=50)

plt.ylim(bottom=0.8*min(avg_r_squ_chatgpt.min(), avg_r_squ_lin_reg.min()), top=1.2*max(avg_r_squ_chatgpt.max(), avg_r_squ_lin_reg.max()))
plt.xticks(results_chatgpt_df['Nb_Years_Train'].unique())

plt.title("F1-score moyen du signe d'Alpha sur les périodes d'entrainement (de 1 à 5 ans)")
plt.legend(loc='upper right')
plt.xlabel("Largeur d'intervalle d'entrainement (années)")
plt.ylabel("F1-score moyen du signe d'Alpha")

# Fit a polynomial (trend line) using numpy.polyfit, adjust degree as needed
degree = 2
coefficients_df1 = np.polyfit(avg_r_squ_chatgpt.index, avg_r_squ_chatgpt.values, degree)
coefficients_df2 = np.polyfit(avg_r_squ_lin_reg.index, avg_r_squ_lin_reg.values, degree)

# Create the trend line using numpy.poly1d
trend_line_df1 = np.poly1d(coefficients_df1)
trend_line_df2 = np.poly1d(coefficients_df2)

# Plot the trend line
plt.plot(avg_r_squ_chatgpt.index, trend_line_df1(avg_r_squ_chatgpt.index), color='green', linestyle='-', alpha=0.3)
plt.plot(avg_r_squ_lin_reg.index, trend_line_df2(avg_r_squ_lin_reg.index), color='blue', linestyle='-', alpha=0.3)
plt.savefig('4a_Avg_Sign_F1_score_Across_Eval_Years.png', dpi=300)
plt.close()
plt.clf()

###### CHART 4b
# Group by 'Eval_Year' and calculate the average 'Sign_F1_score' for each 'Eval_Year'
avg_r_squ_chatgpt = results_chatgpt_df.groupby('Eval_Year')['Sign_F1_score'].mean()
avg_r_squ_lin_reg = results_lin_reg_df.groupby('Eval_Year')['Sign_F1_score'].mean()

plt.scatter(avg_r_squ_chatgpt.index, avg_r_squ_chatgpt.values, label='ChatGPT', color='green', marker='X', s=50)
plt.scatter(avg_r_squ_lin_reg.index, avg_r_squ_lin_reg.values, label='Régression linéaire', color='blue', marker='X', s=50)

plt.ylim(bottom=0.8*min(avg_r_squ_chatgpt.min(), avg_r_squ_lin_reg.min()), top=1.2*max(avg_r_squ_chatgpt.max(), avg_r_squ_lin_reg.max()))
plt.title("F1-score moyen du signe d'Alpha sur chaque année d'évaluation")
plt.legend(loc='upper left')
plt.xlabel("Année d'évaluation")
plt.ylabel("F1-score moyen du signe d'Alpha")
plt.savefig('4b_Avg_Sign_F1_score_Across_Eval_Years.png', dpi=300)
plt.close()
plt.clf()

###### CHART 4c
# Scatter of 'Sign_F1_score' from each experiment of each 'Eval_Year'
merged_df = pd.merge(results_chatgpt_df, results_lin_reg_df, left_index=True, right_index=True, suffixes=('_df1', '_df2'))
plt.scatter(merged_df['Eval_Year_df1'], merged_df['Sign_F1_score_df1'], label='ChatGPT', s=100, c='white', edgecolors='green', linewidth=2)
plt.scatter(merged_df['Eval_Year_df2'], merged_df['Sign_F1_score_df2'], label='Régression linéaire', s=50, c='blue', edgecolors='blue', linewidth=0)

plt.ylim(bottom=-0.05, top=1.25*max(merged_df['Sign_F1_score_df1'].max(), merged_df['Sign_F1_score_df2'].max()))
plt.xlabel("Année d'évaluation")
plt.ylabel("F1-score du signe d'Alpha")
plt.title("F1-score du signe d'Alpha pour chaque instance de test (5 tests par année d'évaluation)", fontsize = 11.5)
plt.legend(loc='upper left')
plt.savefig('4c_Sign_F1_score_of_each_test.png', dpi=300)
plt.close()
plt.clf()
#------------------------------------------------------------------------------------